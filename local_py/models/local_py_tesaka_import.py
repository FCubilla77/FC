# -*- coding: utf-8 -*-

import base64
import io

from odoo import api, fields, models
from odoo.exceptions import UserError


class LocalPyTesakaImport(models.Model):
    _name = 'local_py.tesaka_import'
    _description = 'Importación Tesaka (Retenciones levantadas/anuladas)'
    _order = 'fecha_importacion desc'

    name = fields.Char(string='Referencia', default='Nuevo', copy=False)
    company_id = fields.Many2one(
        'res.company', string='Compañía', required=True, default=lambda self: self.env.company,
    )
    fecha_importacion = fields.Datetime(string='Fecha de Importación', default=fields.Datetime.now, readonly=True)
    usuario_id = fields.Many2one(
        'res.users', string='Usuario', default=lambda self: self.env.user, readonly=True,
    )
    archivo = fields.Binary(string='Archivo Excel (Tesaka)', attachment=True)
    archivo_nombre = fields.Char(string='Nombre del Archivo')
    procesado = fields.Boolean(string='Procesado', default=False, readonly=True)
    linea_ids = fields.One2many('local_py.tesaka_import.linea', 'import_id', string='Líneas')
    total_lineas = fields.Integer(string='Total Líneas', compute='_compute_totales')
    total_exito = fields.Integer(string='Procesadas OK', compute='_compute_totales')
    total_error = fields.Integer(string='Con Error', compute='_compute_totales')

    @api.depends('linea_ids.resultado')
    def _compute_totales(self):
        for imp in self:
            imp.total_lineas = len(imp.linea_ids)
            imp.total_exito = len(imp.linea_ids.filtered(lambda l: l.resultado == 'exito'))
            imp.total_error = len(imp.linea_ids.filtered(lambda l: l.resultado == 'error'))

    @api.model_create_multi
    def create(self, vals_list):
        for vals in vals_list:
            if vals.get('name', 'Nuevo') == 'Nuevo':
                vals['name'] = self.env['ir.sequence'].next_by_code('local_py.tesaka_import') or 'Nuevo'
        return super().create(vals_list)

    # ------------------------------------------------------------------
    # Lectura del Excel
    # ------------------------------------------------------------------
    COLUMNAS = [
        'Tipo', 'Estado', 'Fecha de Anulación', 'Comprobante', 'RUC Informante', 'Informante',
        'Informado - Tipo de identificación', 'RUC Informado', 'Identificación', 'Informado',
        'Control', 'Fecha Emisión', 'Total Renta', 'Comprobante Venta', 'Número Factura Comisión',
        'Timbrado', 'Timbrado Factura Comisión', 'Fecha Recepción', 'Forma de Presentación',
        'Concepto IVA', 'Concepto RENTA', 'Retenido IVA', 'Total Cabezas', 'Total Toneladas',
    ]

    def _leer_filas_excel(self):
        self.ensure_one()
        try:
            from openpyxl import load_workbook
        except ImportError:
            raise UserError('Falta la librería openpyxl en el servidor para leer archivos Excel.')

        wb = load_workbook(io.BytesIO(base64.b64decode(self.archivo)), data_only=True)
        ws = wb.active
        encabezados = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
        idx = {nombre: pos for pos, nombre in enumerate(encabezados) if nombre}

        faltantes = [c for c in ('Estado', 'Comprobante', 'Comprobante Venta') if c not in idx]
        if faltantes:
            raise UserError(
                'El archivo no tiene el formato esperado — faltan las columnas: %s.' % ', '.join(faltantes)
            )

        filas = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not any(v not in (None, '') for v in row):
                continue
            fila = {}
            for nombre, pos in idx.items():
                fila[nombre] = row[pos] if pos < len(row) else None
            filas.append(fila)
        return filas

    def action_procesar(self):
        self.ensure_one()
        if not self.archivo:
            raise UserError('Cargue un archivo Excel antes de procesar.')

        self.linea_ids.unlink()
        filas = self._leer_filas_excel()

        Linea = self.env['local_py.tesaka_import.linea']
        lineas = Linea.browse()
        for fila in filas:
            fecha_emision = fila.get('Fecha Emisión')
            if hasattr(fecha_emision, 'date'):
                fecha_emision = fecha_emision.date()
            fecha_anulacion = fila.get('Fecha de Anulación')
            if hasattr(fecha_anulacion, 'date'):
                fecha_anulacion = fecha_anulacion.date()
            elif not fecha_anulacion:
                fecha_anulacion = False

            lineas |= Linea.create({
                'import_id': self.id,
                'tipo': str(fila.get('Tipo') or ''),
                'estado_dnit': str(fila.get('Estado') or ''),
                'fecha_anulacion_dnit': fecha_anulacion,
                'comprobante': str(fila.get('Comprobante') or ''),
                'ruc_informado': str(fila.get('RUC Informado') or ''),
                'identificacion': str(fila.get('Identificación') or ''),
                'informado': str(fila.get('Informado') or ''),
                'control': str(fila.get('Control') or ''),
                'fecha_emision_dnit': fecha_emision or False,
                'comprobante_venta': str(fila.get('Comprobante Venta') or ''),
                'concepto_iva': str(fila.get('Concepto IVA') or ''),
                'retenido_iva': fila.get('Retenido IVA') or 0.0,
            })

        # Pasada 1: todo lo Aceptado pasa a Levantada primero.
        for linea in lineas:
            if linea.estado_dnit.strip().lower() == 'aceptado':
                linea._procesar_aceptada()

        # Pasada 2: recién ahí, lo Anulado — así, si el mismo archivo trae
        # una Retención que fue Aceptada y luego Anulada, siempre queda
        # bien secuenciado (Pendiente -> Levantada -> Anulada), sin
        # importar el orden de las filas en el Excel.
        for linea in lineas:
            estado = linea.estado_dnit.strip().lower()
            if estado and estado != 'aceptado':
                linea._procesar_anulada()

        self.procesado = True
        return True
